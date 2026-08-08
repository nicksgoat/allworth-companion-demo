"""Tests for the Pipeline Review blueprint.

Covers:
- Row → prospect conversion and JSON-safety helpers
- API endpoints (weeks, filters, snapshot, trend, movement) with a mocked Synapse cursor
- XLSX export returns a valid workbook

Run from the backend/ directory:

    python -m pytest tests/test_pipeline_review.py -v
"""
from __future__ import annotations

import io
import os
from unittest.mock import MagicMock, patch

import pytest

# Disable auth so tests can hit routes without a JWT
os.environ["AUTH_DISABLE"] = "1"

from pipeline_review.routes import (  # noqa: E402
    _cache_clear,
    _json_safe,
    _row_to_prospect,
)


@pytest.fixture(autouse=True)
def clear_cache():
    _cache_clear()
    yield
    _cache_clear()


@pytest.fixture()
def app():
    from flask import Flask
    from pipeline_review.routes import bp

    app = Flask(__name__)
    app.testing = True
    app.register_blueprint(bp, url_prefix="/pipeline-review")
    return app


@pytest.fixture()
def client(app):
    return app.test_client()


def _mock_conn(cursor):
    conn = MagicMock()
    conn.cursor.return_value = cursor
    return conn


# ---------------------------------------------------------------------------
# Unit: helpers
# ---------------------------------------------------------------------------

class TestHelpers:
    def test_json_safe_replaces_nan_inf(self):
        assert _json_safe(float("nan")) is None
        assert _json_safe(float("inf")) is None
        assert _json_safe({"a": float("nan"), "b": [1.0, float("-inf")]}) == {
            "a": None, "b": [1.0, None],
        }
        assert _json_safe(3.5) == 3.5

    def test_row_to_prospect_types(self):
        row = (
            "abc123", "2026-W28", "2026-07-13", "00Q1", "Jane Doe", 3_500_000.0,
            "6 - Proposal Delivered", 42, 60, 95, "Whale prospect", "Advisor Driven",
            "Adv Name", "RVP Name", "West", "2026-08-01", "2026-07-10", "2026-07-20", True,
        )
        p = _row_to_prospect(row)
        assert p["id"] == "abc123"
        assert p["paum"] == 3_500_000.0
        assert isinstance(p["days_in_stage"], int) and p["days_in_stage"] == 42
        assert isinstance(p["avg_dwell"], int) and p["avg_dwell"] == 60
        assert isinstance(p["score"], int) and p["score"] == 95
        assert p["reasons"] == "Whale prospect"
        assert p["expected_close_date"] == "2026-08-01"
        assert p["next_activity_date"] == "2026-07-20"
        assert p["was_stale"] is True
        assert p["region"] == "West"
        assert p["sf_url"].endswith("/lead/00Q1/view")

    def test_row_to_prospect_handles_nulls(self):
        row = ("id1", "2026-W28", None, "00Q2", None, None, None, None, None, None,
               None, None, None, None, None, None, None, None, None)
        p = _row_to_prospect(row)
        assert p["paum"] == 0.0
        assert p["days_in_stage"] == 0
        assert p["avg_dwell"] == 0
        assert p["score"] == 0
        assert p["name"] == ""
        assert p["reasons"] == ""
        assert p["expected_close_date"] is None
        assert p["was_stale"] is False


# ---------------------------------------------------------------------------
# API endpoints
# ---------------------------------------------------------------------------

class TestEndpoints:
    def test_weeks(self, client):
        cursor = MagicMock()
        cursor.fetchall.return_value = [("2026-W28", "2026-07-13"), ("2026-W27", "2026-07-06")]
        with patch("pipeline_review.routes._get_db_connection", return_value=_mock_conn(cursor)):
            rv = client.get("/pipeline-review/api/weeks")
        assert rv.status_code == 200
        body = rv.get_json()
        assert body["success"] is True
        assert body["data"] == [
            {"snapshot_week": "2026-W28", "report_date": "2026-07-13"},
            {"snapshot_week": "2026-W27", "report_date": "2026-07-06"},
        ]

    def test_filters(self, client):
        cursor = MagicMock()
        # region, channel, stage, advisor distinct calls in order
        cursor.fetchall.side_effect = [
            [("West",), ("East",)],
            [("Advisor Driven",)],
            [("6 - Proposal Delivered",)],
            [("Adv Name",)],
        ]
        with patch("pipeline_review.routes._get_db_connection", return_value=_mock_conn(cursor)):
            rv = client.get("/pipeline-review/api/filters")
        assert rv.status_code == 200
        data = rv.get_json()["data"]
        assert data["regions"] == ["West", "East"]
        assert data["channels"] == ["Advisor Driven"]
        assert data["stages"] == ["6 - Proposal Delivered"]
        assert data["advisors"] == ["Adv Name"]

    def test_snapshot_uses_latest_when_no_week(self, client):
        cursor = MagicMock()
        # 1) MAX(snapshot_week); 2) detail rows; 3) summary row
        cursor.fetchone.side_effect = [
            ("2026-W28",),
            ("2026-W28", "2026-07-13", 2, 5_000_000.0, 4_000_000.0, 3_000_000.0,
             1, 1, 2_000_000.0, 1, 2_000_000.0),
        ]
        detail_row = (
            "abc", "2026-W28", "2026-07-13", "00Q1", "Jane", 3_000_000.0,
            "6 - Proposal Delivered", 10, 45, 90, "High-value", "Advisor Driven",
            "Adv", "RVP", "West", "2026-08-01", "2026-07-10", None, False,
        )
        cursor.fetchall.return_value = [detail_row]
        with patch("pipeline_review.routes._get_db_connection", return_value=_mock_conn(cursor)):
            rv = client.get("/pipeline-review/api/snapshot")
        assert rv.status_code == 200
        data = rv.get_json()["data"]
        assert data["week"] == "2026-W28"
        assert len(data["prospects"]) == 1
        assert data["summary"]["total_prospects"] == 2
        assert data["summary"]["weighted_pipeline"] == 4_000_000.0
        assert data["summary"]["closing_next_30_count"] == 1

    def test_snapshot_empty_when_no_data(self, client):
        cursor = MagicMock()
        cursor.fetchone.side_effect = [(None,)]  # MAX returns NULL
        with patch("pipeline_review.routes._get_db_connection", return_value=_mock_conn(cursor)):
            rv = client.get("/pipeline-review/api/snapshot")
        assert rv.status_code == 200
        data = rv.get_json()["data"]
        assert data["week"] is None
        assert data["prospects"] == []

    def test_trend(self, client):
        cursor = MagicMock()
        cursor.fetchall.return_value = [
            ("2026-W27", "2026-07-06", 3, 4_000_000.0, 2_000_000.0, 1_000_000.0, 0, 0, 0.0, 0, 0.0),
            ("2026-W28", "2026-07-13", 4, 5_000_000.0, 3_000_000.0, 2_000_000.0, 1, 1, 2_000_000.0, 1, 2_000_000.0),
        ]
        with patch("pipeline_review.routes._get_db_connection", return_value=_mock_conn(cursor)):
            rv = client.get("/pipeline-review/api/trend")
        assert rv.status_code == 200
        series = rv.get_json()["data"]
        assert len(series) == 2
        assert series[1]["total_prospects"] == 4
        assert series[1]["weighted_pipeline"] == 3_000_000.0

    def test_movement_new_and_dropped(self, client):
        cursor = MagicMock()
        # MAX week, then prior-week MAX
        cursor.fetchone.side_effect = [("2026-W28",), ("2026-W27",)]
        # current rows, then prior rows
        cursor.fetchall.side_effect = [
            [("L1", "Jane", "6 - Proposal Delivered", 90, 3_000_000.0),
             ("L2", "Bob", "5 - Discovery", 40, 2_500_000.0)],
            [("L1", "Jane", "5 - Discovery", 60, 3_000_000.0),
             ("L3", "Sue", "7 - Verbal Commitment Received", 20, 8_000_000.0)],
        ]
        with patch("pipeline_review.routes._get_db_connection", return_value=_mock_conn(cursor)):
            rv = client.get("/pipeline-review/api/movement")
        assert rv.status_code == 200
        data = rv.get_json()["data"]
        assert data["prior_week"] == "2026-W27"
        assert [p["lead_id"] for p in data["new"]] == ["L2"]
        assert [p["lead_id"] for p in data["dropped"]] == ["L3"]
        assert [p["lead_id"] for p in data["advanced"]] == ["L1"]

    def test_closed_returns_rows(self, client):
        cursor = MagicMock()
        cursor.fetchone.side_effect = [("2026-W28",)]  # latest week
        cursor.fetchall.return_value = [
            ("L9", "Won Co", 6_000_000.0, "CRP", "8 - Onboarding", "Adv", "RVP",
             "West", "2026-07-11", 2),
        ]
        with patch("pipeline_review.routes._get_db_connection", return_value=_mock_conn(cursor)):
            rv = client.get("/pipeline-review/api/closed")
        assert rv.status_code == 200
        data = rv.get_json()["data"]
        assert data["week"] == "2026-W28"
        assert len(data["closed"]) == 1
        c = data["closed"][0]
        assert c["paum"] == 6_000_000.0
        assert c["channel"] == "CRP"
        assert c["sf_url"].endswith("/lead/L9/view")

    def test_progress_returns_rows(self, client):
        cursor = MagicMock()
        cursor.fetchone.side_effect = [("2026-W28",)]  # latest week
        cursor.fetchall.return_value = [
            # won (paum 5M), advanced (paum 3M), stable (paum 1M), dropped (paum 2.5M)
            ("L1", "Won Co", 5_000_000.0, "CRP", "Adv", "RVP", "West",
             "8 - Onboarding", 90, None, None, "won", "$5.0M converted"),
            ("L2", "Move Co", 3_000_000.0, "Advisor Driven", "Adv2", "RVP2", "East",
             "5 - Discovery", 60, "7 - Verbal Commitment Received", 85, "advanced",
             "Discovery -> Verbal"),
            ("L3", "Stay Co", 1_000_000.0, "Paid Leads", "Adv3", "RVP3", "West",
             "5 - Discovery", 40, "5 - Discovery", 40, "stable", ""),
            ("L4", "Gone Co", 2_500_000.0, "Media Driven", "Adv4", "RVP4", "East",
             "5 - Discovery", 30, None, None, "dropped", "Stage changed to 'Nurture'"),
        ]
        with patch("pipeline_review.routes._get_db_connection", return_value=_mock_conn(cursor)):
            rv = client.get("/pipeline-review/api/progress")
        assert rv.status_code == 200
        data = rv.get_json()["data"]
        assert data["week"] == "2026-W28"
        # sorted by status rank then -paum: won, advanced, stable, dropped
        assert [r["status"] for r in data["rows"]] == ["won", "advanced", "stable", "dropped"]
        assert data["totals"]["prior_prospects"] == 4
        assert data["totals"]["won"] == {"count": 1, "paum": 5_000_000.0}
        assert data["totals"]["advanced"]["count"] == 1
        assert data["totals"]["dropped"]["paum"] == 2_500_000.0
        won = data["rows"][0]
        assert won["score_prev"] == 90 and won["score_now"] is None
        assert won["sf_url"].endswith("/lead/L1/view")

    def test_progress_empty_when_no_data(self, client):
        cursor = MagicMock()
        cursor.fetchone.side_effect = [(None,)]  # MAX returns NULL
        with patch("pipeline_review.routes._get_db_connection", return_value=_mock_conn(cursor)):
            rv = client.get("/pipeline-review/api/progress")
        assert rv.status_code == 200
        data = rv.get_json()["data"]
        assert data["week"] is None
        assert data["rows"] == []

    def test_export_excel_returns_workbook(self, client):
        payload = {
            "week": "2026-W28",
            "prospects": [
                {"name": "Jane", "paum": 3_000_000.0, "stage": "6 - Proposal Delivered",
                 "days_in_stage": 10, "score": 90, "channel": "Advisor Driven",
                 "advisor_name": "Adv", "sales_person": "RVP", "region": "West",
                 "was_stale": True, "worked": False},
            ],
        }
        rv = client.post("/pipeline-review/api/export-excel", json=payload)
        assert rv.status_code == 200
        assert rv.mimetype == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # openpyxl can reopen the produced bytes
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(rv.data))
        ws = wb.active
        assert ws["A1"].value == "Prospect"
        assert ws["A2"].value == "Jane"

    def test_cache_clear(self, client):
        rv = client.post("/pipeline-review/api/cache-clear")
        assert rv.status_code == 200
        assert rv.get_json()["success"] is True
