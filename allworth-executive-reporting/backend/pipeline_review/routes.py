"""Weekly Pipeline Review routes — read-only snapshot history, focus list, trend, XLSX export.

Data source: physical [tho].[Pipeline_Review_Snapshot] (one row per snapshot_week + lead)
and [tho].[Pipeline_Review_Summary] (one row per snapshot_week), both filled weekly by the
Synapse Spark notebook ``pipeline_review_snapshot_build``. This blueprint is READ-ONLY — it
never writes to the warehouse. Per-user "worked / checked off" state lives in the browser
(localStorage), not here, because the snapshot table is rebuilt every week.
"""

from __future__ import annotations

import io
import math
import time
from pathlib import Path
from typing import Any

from flask import Blueprint, jsonify, request, send_file, send_from_directory

bp = Blueprint(
    "pipeline_review",
    __name__,
    template_folder="templates",
)


# ─── JSON safety ─────────────────────────────────────────────────────────────

def _json_safe(value: Any) -> Any:
    """Replace NaN/Infinity (which browsers' JSON.parse rejects) with None."""
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, dict):
        return {k: _json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    return value


# ─── DB + cache (mirrors fee_calculator conventions) ─────────────────────────

def _get_db_connection():
    """Synapse connection via the shared pool in app.py (no hardcoded creds)."""
    from app import get_database_connection
    return get_database_connection()


_cache: dict[str, tuple[float, Any]] = {}
_CACHE_TTL_SECONDS = 1800  # 30 min — the snapshot table only changes once a week


def _cache_get(key: str) -> Any | None:
    entry = _cache.get(key)
    if entry and (time.time() - entry[0]) < _CACHE_TTL_SECONDS:
        return entry[1]
    return None


def _cache_set(key: str, value: Any) -> None:
    _cache[key] = (time.time(), value)


def _cache_clear() -> None:
    _cache.clear()


# ─── Column order for the detail table (single source of truth) ──────────────

_DETAIL_COLUMNS = [
    "id", "snapshot_week", "report_date", "lead_id", "name", "paum", "stage",
    "days_in_stage", "avg_dwell", "score", "reasons", "channel", "advisor_name",
    "sales_person", "region", "expected_close_date", "last_activity_date",
    "next_activity_date", "was_stale",
]

_DETAIL_SELECT = (
    "id, snapshot_week, CONVERT(VARCHAR(10), report_date, 23) AS report_date, "
    "lead_id, name, paum, stage, days_in_stage, avg_dwell, score, reasons, channel, "
    "advisor_name, sales_person, region, "
    "CONVERT(VARCHAR(10), expected_close_date, 23) AS expected_close_date, "
    "CONVERT(VARCHAR(10), last_activity_date, 23) AS last_activity_date, "
    "CONVERT(VARCHAR(10), next_activity_date, 23) AS next_activity_date, "
    "was_stale"
)


def _row_to_prospect(row) -> dict[str, Any]:
    d = dict(zip(_DETAIL_COLUMNS, row))
    d["paum"] = float(d["paum"]) if d["paum"] is not None else 0.0
    d["days_in_stage"] = int(d["days_in_stage"]) if d["days_in_stage"] is not None else 0
    d["avg_dwell"] = int(d["avg_dwell"]) if d["avg_dwell"] is not None else 0
    d["score"] = int(d["score"]) if d["score"] is not None else 0
    d["was_stale"] = bool(d["was_stale"])
    for k in ("expected_close_date", "last_activity_date", "next_activity_date"):
        d[k] = str(d[k]) if d[k] is not None else None
    for k in ("name", "stage", "reasons", "channel", "advisor_name", "sales_person", "region"):
        d[k] = str(d[k]) if d[k] is not None else ""
    d["sf_url"] = (
        f"https://allworth.lightning.force.com/lightning/r/lead/{d['lead_id']}/view"
        if d.get("lead_id") else ""
    )
    return d


def _latest_week(cursor) -> str | None:
    cursor.execute("SELECT MAX(snapshot_week) FROM [tho].[Pipeline_Review_Snapshot]")
    row = cursor.fetchone()
    return str(row[0]) if row and row[0] is not None else None


def _summary_row_to_dict(r) -> dict[str, Any]:
    return {
        "snapshot_week": str(r[0]),
        "report_date": str(r[1]) if r[1] is not None else None,
        "total_prospects": int(r[2]) if r[2] is not None else 0,
        "total_paum": float(r[3]) if r[3] is not None else 0.0,
        "weighted_pipeline": float(r[4]) if r[4] is not None else 0.0,
        "verbal_onboarding_paum": float(r[5]) if r[5] is not None else 0.0,
        "verbal_onboarding_count": int(r[6]) if r[6] is not None else 0,
        "closing_next_30_count": int(r[7]) if r[7] is not None else 0,
        "closing_next_30_paum": float(r[8]) if r[8] is not None else 0.0,
        "conversions_count": int(r[9]) if r[9] is not None else 0,
        "conversions_paum": float(r[10]) if r[10] is not None else 0.0,
    }


# ─── Routes ──────────────────────────────────────────────────────────────────

@bp.route("/")
def index():
    """Serve the Pipeline Review SPA."""
    return send_from_directory(str(Path(__file__).parent / "templates"), "index.html")


@bp.route("/api/cache-clear", methods=["POST"])
def clear_cache():
    """Force fresh Synapse reads on the next request."""
    _cache_clear()
    return jsonify({"success": True, "message": "Cache cleared"})


@bp.route("/api/weeks", methods=["GET"])
def get_weeks():
    """Return available snapshot weeks (newest first) with each week's report date."""
    cached = _cache_get("weeks")
    if cached is not None:
        return jsonify({"success": True, "data": cached})
    try:
        conn = _get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT snapshot_week, CONVERT(VARCHAR(10), MAX(report_date), 23) AS report_date "
            "FROM [tho].[Pipeline_Review_Snapshot] "
            "GROUP BY snapshot_week ORDER BY snapshot_week DESC"
        )
        weeks = [
            {"snapshot_week": str(r[0]), "report_date": str(r[1]) if r[1] is not None else None}
            for r in cursor.fetchall()
        ]
        cursor.close()
        _cache_set("weeks", weeks)
        return jsonify({"success": True, "data": weeks})
    except Exception as e:  # pragma: no cover - defensive
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/filters", methods=["GET"])
def get_filters():
    """Distinct region / channel / stage / advisor values for the filter controls."""
    cached = _cache_get("filters")
    if cached is not None:
        return jsonify({"success": True, "data": cached})
    try:
        conn = _get_db_connection()
        cursor = conn.cursor()

        def _distinct(col: str) -> list[str]:
            cursor.execute(
                f"SELECT DISTINCT {col} FROM [tho].[Pipeline_Review_Snapshot] "
                f"WHERE {col} IS NOT NULL AND {col} <> '' ORDER BY {col}"
            )
            return [str(r[0]) for r in cursor.fetchall()]

        data = {
            "regions": _distinct("region"),
            "channels": _distinct("channel"),
            "stages": _distinct("stage"),
            "advisors": _distinct("advisor_name"),
        }
        cursor.close()
        _cache_set("filters", data)
        return jsonify({"success": True, "data": data})
    except Exception as e:  # pragma: no cover - defensive
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/snapshot", methods=["GET"])
def get_snapshot():
    """Return the focus list + summary for one week (defaults to the latest).

    Query params: week=<snapshot_week> (optional).
    """
    week = request.args.get("week", "").strip()
    cache_key = f"snapshot:{week or 'latest'}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return jsonify({"success": True, "data": cached})
    try:
        conn = _get_db_connection()
        cursor = conn.cursor()

        if not week:
            week = _latest_week(cursor)
        if not week:
            cursor.close()
            return jsonify({"success": True, "data": {"week": None, "prospects": [], "summary": None}})

        cursor.execute(
            f"SELECT {_DETAIL_SELECT} FROM [tho].[Pipeline_Review_Snapshot] "
            "WHERE snapshot_week = ? ORDER BY score DESC, paum DESC",
            week,
        )
        prospects = [_row_to_prospect(r) for r in cursor.fetchall()]

        cursor.execute(
            "SELECT snapshot_week, CONVERT(VARCHAR(10), report_date, 23), total_prospects, "
            "total_paum, weighted_pipeline, verbal_onboarding_paum, verbal_onboarding_count, "
            "closing_next_30_count, closing_next_30_paum, conversions_count, conversions_paum "
            "FROM [tho].[Pipeline_Review_Summary] WHERE snapshot_week = ?",
            week,
        )
        srow = cursor.fetchone()
        summary = None
        if srow:
            summary = _summary_row_to_dict(srow)
        cursor.close()

        data = _json_safe({"week": week, "prospects": prospects, "summary": summary})
        _cache_set(cache_key, data)
        return jsonify({"success": True, "data": data})
    except Exception as e:  # pragma: no cover - defensive
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/trend", methods=["GET"])
def get_trend():
    """Return the weekly summary series (all weeks) for the trend charts."""
    cached = _cache_get("trend")
    if cached is not None:
        return jsonify({"success": True, "data": cached})
    try:
        conn = _get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT snapshot_week, CONVERT(VARCHAR(10), report_date, 23), total_prospects, "
            "total_paum, weighted_pipeline, verbal_onboarding_paum, verbal_onboarding_count, "
            "closing_next_30_count, closing_next_30_paum, conversions_count, conversions_paum "
            "FROM [tho].[Pipeline_Review_Summary] ORDER BY snapshot_week ASC"
        )
        series = [_summary_row_to_dict(r) for r in cursor.fetchall()]
        cursor.close()
        data = _json_safe(series)
        _cache_set("trend", data)
        return jsonify({"success": True, "data": data})
    except Exception as e:  # pragma: no cover - defensive
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/movement", methods=["GET"])
def get_movement():
    """Week-over-week movement vs the prior snapshot: new / advanced / dropped / stable.

    Query params: week=<snapshot_week> (optional, defaults to latest).
    """
    week = request.args.get("week", "").strip()
    cache_key = f"movement:{week or 'latest'}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return jsonify({"success": True, "data": cached})
    try:
        conn = _get_db_connection()
        cursor = conn.cursor()

        if not week:
            week = _latest_week(cursor)
        if not week:
            cursor.close()
            return jsonify({"success": True, "data": {"week": None, "prior_week": None,
                                                      "new": [], "advanced": [], "dropped": []}})

        cursor.execute(
            "SELECT MAX(snapshot_week) FROM [tho].[Pipeline_Review_Snapshot] "
            "WHERE snapshot_week < ?",
            week,
        )
        prow = cursor.fetchone()
        prior_week = str(prow[0]) if prow and prow[0] is not None else None

        cursor.execute(
            "SELECT lead_id, name, stage, score, paum FROM [tho].[Pipeline_Review_Snapshot] "
            "WHERE snapshot_week = ?",
            week,
        )
        current = {str(r[0]): {"lead_id": str(r[0]), "name": str(r[1] or ""), "stage": str(r[2] or ""),
                               "score": int(r[3] or 0), "paum": float(r[4] or 0)} for r in cursor.fetchall()}

        prior: dict[str, dict[str, Any]] = {}
        if prior_week:
            cursor.execute(
                "SELECT lead_id, name, stage, score, paum FROM [tho].[Pipeline_Review_Snapshot] "
                "WHERE snapshot_week = ?",
                prior_week,
            )
            prior = {str(r[0]): {"lead_id": str(r[0]), "name": str(r[1] or ""), "stage": str(r[2] or ""),
                                 "score": int(r[3] or 0), "paum": float(r[4] or 0)} for r in cursor.fetchall()}
        cursor.close()

        new_leads = [v for k, v in current.items() if k not in prior]
        dropped = [v for k, v in prior.items() if k not in current]
        advanced = []
        for k, cur in current.items():
            if k in prior and cur["stage"] != prior[k]["stage"]:
                advanced.append({**cur, "prior_stage": prior[k]["stage"]})

        data = _json_safe({
            "week": week,
            "prior_week": prior_week,
            "new": new_leads,
            "advanced": advanced,
            "dropped": dropped,
        })
        _cache_set(cache_key, data)
        return jsonify({"success": True, "data": data})
    except Exception as e:  # pragma: no cover - defensive
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/closed", methods=["GET"])
def get_closed():
    """Return the leads closed-won in the last 7 days for one week (defaults to latest)."""
    week = request.args.get("week", "").strip()
    cache_key = f"closed:{week or 'latest'}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return jsonify({"success": True, "data": cached})
    try:
        conn = _get_db_connection()
        cursor = conn.cursor()

        if not week:
            week = _latest_week(cursor)
        if not week:
            cursor.close()
            return jsonify({"success": True, "data": {"week": None, "closed": []}})

        cursor.execute(
            "SELECT lead_id, name, paum, channel, stage_prev, advisor_name, sales_person, "
            "region, CONVERT(VARCHAR(10), close_date, 23), days_closed "
            "FROM [tho].[Pipeline_Review_Closed] WHERE snapshot_week = ? "
            "ORDER BY paum DESC",
            week,
        )
        closed = [
            {
                "lead_id": str(r[0] or ""),
                "name": str(r[1] or ""),
                "paum": float(r[2]) if r[2] is not None else 0.0,
                "channel": str(r[3] or ""),
                "stage_prev": str(r[4] or ""),
                "advisor_name": str(r[5] or ""),
                "sales_person": str(r[6] or ""),
                "region": str(r[7] or ""),
                "close_date": str(r[8]) if r[8] is not None else None,
                "days_closed": int(r[9]) if r[9] is not None else 0,
                "sf_url": (
                    f"https://allworth.lightning.force.com/lightning/r/lead/{r[0]}/view"
                    if r[0] else ""
                ),
            }
            for r in cursor.fetchall()
        ]
        cursor.close()

        data = _json_safe({"week": week, "closed": closed})
        _cache_set(cache_key, data)
        return jsonify({"success": True, "data": data})
    except Exception as e:  # pragma: no cover - defensive
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/progress", methods=["GET"])
def get_progress():
    """Week-over-week progress for one week: prior-week leads classified as
    won / advanced / re-engaged / stable / dropped (defaults to the latest week).

    Computed at snapshot-build time in Synapse (the "won" status needs live
    Current_Client / closed-won state the app cannot derive from two snapshots).
    """
    week = request.args.get("week", "").strip()
    cache_key = f"progress:{week or 'latest'}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return jsonify({"success": True, "data": cached})
    try:
        conn = _get_db_connection()
        cursor = conn.cursor()

        if not week:
            week = _latest_week(cursor)
        if not week:
            cursor.close()
            return jsonify({"success": True, "data": {"week": None, "rows": [], "totals": {}}})

        cursor.execute(
            "SELECT lead_id, name, paum, channel, advisor_name, sales_person, region, "
            "stage_prev, score_prev, stage_now, score_now, progress_status, note "
            "FROM [tho].[Pipeline_Review_Progress] WHERE snapshot_week = ? "
            "ORDER BY paum DESC",
            week,
        )
        _status_rank = {"won": 0, "advanced": 1, "re-engaged": 2, "stable": 3, "dropped": 4}
        rows = [
            {
                "lead_id": str(r[0] or ""),
                "name": str(r[1] or ""),
                "paum": float(r[2]) if r[2] is not None else 0.0,
                "channel": str(r[3] or ""),
                "advisor_name": str(r[4] or ""),
                "sales_person": str(r[5] or ""),
                "region": str(r[6] or ""),
                "stage_prev": str(r[7] or ""),
                "score_prev": int(r[8]) if r[8] is not None else None,
                "stage_now": str(r[9]) if r[9] is not None else None,
                "score_now": int(r[10]) if r[10] is not None else None,
                "status": str(r[11] or ""),
                "note": str(r[12] or ""),
                "sf_url": (
                    f"https://allworth.lightning.force.com/lightning/r/lead/{r[0]}/view"
                    if r[0] else ""
                ),
            }
            for r in cursor.fetchall()
        ]
        cursor.close()

        rows.sort(key=lambda x: (_status_rank.get(x["status"], 9), -x["paum"]))

        def _agg(status: str) -> dict[str, Any]:
            sel = [x for x in rows if x["status"] == status]
            return {"count": len(sel), "paum": sum(x["paum"] for x in sel)}

        totals = {
            "prior_prospects": len(rows),
            "won": _agg("won"),
            "advanced": _agg("advanced"),
            "re_engaged": _agg("re-engaged"),
            "stable": _agg("stable"),
            "dropped": _agg("dropped"),
        }

        data = _json_safe({"week": week, "rows": rows, "totals": totals})
        _cache_set(cache_key, data)
        return jsonify({"success": True, "data": data})
    except Exception as e:  # pragma: no cover - defensive
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/export-excel", methods=["POST"])
def export_excel():
    """Export the provided focus-list rows to a formatted XLSX.

    Body: { "week": "2026-W28", "prospects": [ {...}, ... ] }
    The frontend posts the currently-filtered rows so the export honors filters
    and includes any per-user "worked" status.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    data = request.get_json(force=True) or {}
    week = str(data.get("week") or "").strip()
    prospects = data.get("prospects") or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Focus List"

    headers = [
        ("name", "Prospect"),
        ("paum", "PAUM"),
        ("stage", "Stage"),
        ("days_in_stage", "Days in Stage"),
        ("score", "Score"),
        ("channel", "Channel"),
        ("advisor_name", "Advisor"),
        ("sales_person", "Sales Person"),
        ("region", "Region"),
        ("was_stale", "At Risk"),
        ("worked", "Worked"),
    ]

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    for c, (_key, label) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r, p in enumerate(prospects, start=2):
        for c, (key, _label) in enumerate(headers, start=1):
            val = p.get(key)
            if key == "was_stale":
                val = "Yes" if val else ""
            elif key == "worked":
                val = "Yes" if val else ""
            ws.cell(row=r, column=c, value=val)
        ws.cell(row=r, column=2).number_format = "$#,##0"

    widths = [34, 16, 26, 13, 8, 16, 24, 22, 12, 9, 9]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"pipeline_review_{week or 'latest'}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )
