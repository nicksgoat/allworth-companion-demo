"""Fee Calculator routes — tiered fee computation + household lookup + CSV upload."""

from __future__ import annotations

import csv
import io
import math
import os
import time
from typing import Any

from flask import Blueprint, jsonify, request, send_file, send_from_directory
from pathlib import Path

bp = Blueprint(
    "fee_calculator",
    __name__,
    template_folder="templates",
)


def _json_safe(value: Any) -> Any:
    """Recursively replace non-JSON-safe values so jsonify emits valid JSON.

    Flask's default encoder writes bare ``NaN``/``Infinity`` tokens which the
    browser's ``JSON.parse`` rejects. Pandas string columns aggregated with
    ``"first"`` can also surface float ``NaN`` for blank cells. Convert those
    to ``None`` (and NaN inside strings is impossible, so only floats matter).
    """
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, dict):
        return {k: _json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    return value

# ─── Fee Schedule Definitions (from Excel model) ────────────────────────────

# GM Schedule New (Min Fee)
GM_SCHEDULE_NEW = {
    "name": "GM Schedule New (Min Fee)",
    "min_quarterly_fee": 2500,
    "tiers": [
        {"from": 0, "to": 250000, "rate": 0.0150},
        {"from": 250001, "to": 750000, "rate": 0.0125},
        {"from": 750001, "to": 1500000, "rate": 0.0100},
        {"from": 1500001, "to": 3000000, "rate": 0.0080},
        {"from": 3000001, "to": 5000000, "rate": 0.0070},
        {"from": 5000001, "to": 7500000, "rate": 0.0060},
        {"from": 7500001, "to": 10000000, "rate": 0.0040},
        {"from": 10000001, "to": 50000000, "rate": 0.0035},
        {"from": 50000001, "to": None, "rate": 0.0030},
    ],
}

# New Airline Clients
AIRLINE_SCHEDULE = {
    "name": "New Airline Clients",
    "min_quarterly_fee": 0,
    "tiers": [
        {"from": 0, "to": 500000, "rate": 0.0120},
        {"from": 500001, "to": 1000000, "rate": 0.0110},
        {"from": 1000001, "to": 1500000, "rate": 0.0100},
        {"from": 1500001, "to": 2000000, "rate": 0.0090},
        {"from": 2000001, "to": None, "rate": 0.0070},
    ],
}

# Repricing Process Fee Schedules (5 tiers)
REPRICING_SCHEDULES = {
    "Silver (1)": {
        "name": "Repricing - Silver (1)",
        "min_quarterly_fee": 0,
        "tiers": [
            {"from": 0, "to": 100000, "rate": 0.0200},
            {"from": 100001, "to": 200000, "rate": 0.0150},
            {"from": 200001, "to": 1000000, "rate": 0.0130},
            {"from": 1000001, "to": 2500000, "rate": 0.0100},
            {"from": 2500001, "to": 5000000, "rate": 0.0080},
            {"from": 5000001, "to": None, "rate": 0.0070},
        ],
    },
    "Gold (2)": {
        "name": "Repricing - Gold (2)",
        "min_quarterly_fee": 0,
        "tiers": [
            {"from": 0, "to": 100000, "rate": 0.0140},
            {"from": 100001, "to": 200000, "rate": 0.0140},
            {"from": 200001, "to": 1000000, "rate": 0.0120},
            {"from": 1000001, "to": 2500000, "rate": 0.0100},
            {"from": 2500001, "to": 5000000, "rate": 0.0080},
            {"from": 5000001, "to": None, "rate": 0.0070},
        ],
    },
    "Platinum (3)": {
        "name": "Repricing - Platinum (3)",
        "min_quarterly_fee": 0,
        "tiers": [
            {"from": 0, "to": 100000, "rate": 0.0130},
            {"from": 100001, "to": 200000, "rate": 0.0130},
            {"from": 200001, "to": 1000000, "rate": 0.0110},
            {"from": 1000001, "to": 2500000, "rate": 0.0100},
            {"from": 2500001, "to": 5000000, "rate": 0.0080},
            {"from": 5000001, "to": None, "rate": 0.0070},
        ],
    },
    "Diamond (4)": {
        "name": "Repricing - Diamond (4)",
        "min_quarterly_fee": 0,
        "tiers": [
            {"from": 0, "to": 100000, "rate": 0.0120},
            {"from": 100001, "to": 200000, "rate": 0.0120},
            {"from": 200001, "to": 1000000, "rate": 0.0100},
            {"from": 1000001, "to": 2500000, "rate": 0.0095},
            {"from": 2500001, "to": 5000000, "rate": 0.0085},
            {"from": 5000001, "to": None, "rate": 0.0075},
        ],
    },
    "Elite (5)": {
        "name": "Repricing - Elite (5)",
        "min_quarterly_fee": 0,
        "tiers": [
            {"from": 0, "to": 100000, "rate": 0.0095},
            {"from": 100001, "to": 200000, "rate": 0.0095},
            {"from": 200001, "to": 1000000, "rate": 0.0095},
            {"from": 1000001, "to": 2500000, "rate": 0.0095},
            {"from": 2500001, "to": 5000000, "rate": 0.0085},
            {"from": 5000001, "to": None, "rate": 0.0075},
        ],
    },
}

# Fixed Rate Schedules (flat rate across all AUM)
FIXED_150 = {"name": "1.50% Fixed", "min_quarterly_fee": 0, "tiers": [{"from": 0, "to": None, "rate": 0.0150}]}
FIXED_145 = {"name": "1.45% Fixed", "min_quarterly_fee": 0, "tiers": [{"from": 0, "to": None, "rate": 0.0145}]}
FIXED_140 = {"name": "1.40% Fixed", "min_quarterly_fee": 0, "tiers": [{"from": 0, "to": None, "rate": 0.0140}]}
FIXED_135 = {"name": "1.35% Fixed", "min_quarterly_fee": 0, "tiers": [{"from": 0, "to": None, "rate": 0.0135}]}
FIXED_130 = {"name": "1.30% Fixed", "min_quarterly_fee": 0, "tiers": [{"from": 0, "to": None, "rate": 0.0130}]}
FIXED_125 = {"name": "1.25% Fixed", "min_quarterly_fee": 0, "tiers": [{"from": 0, "to": None, "rate": 0.0125}]}
FIXED_120 = {"name": "1.20% Fixed", "min_quarterly_fee": 0, "tiers": [{"from": 0, "to": None, "rate": 0.0120}]}
FIXED_115 = {"name": "1.15% Fixed", "min_quarterly_fee": 0, "tiers": [{"from": 0, "to": None, "rate": 0.0115}]}
FIXED_110 = {"name": "1.10% Fixed", "min_quarterly_fee": 0, "tiers": [{"from": 0, "to": None, "rate": 0.0110}]}
FIXED_105 = {"name": "1.05% Fixed", "min_quarterly_fee": 0, "tiers": [{"from": 0, "to": None, "rate": 0.0105}]}
FIXED_100 = {"name": "1.00% Fixed", "min_quarterly_fee": 0, "tiers": [{"from": 0, "to": None, "rate": 0.0100}]}
FIXED_095 = {"name": "0.95% Fixed", "min_quarterly_fee": 0, "tiers": [{"from": 0, "to": None, "rate": 0.0095}]}
FIXED_090 = {"name": "0.90% Fixed", "min_quarterly_fee": 0, "tiers": [{"from": 0, "to": None, "rate": 0.0090}]}
FIXED_085 = {"name": "0.85% Fixed", "min_quarterly_fee": 0, "tiers": [{"from": 0, "to": None, "rate": 0.0085}]}
FIXED_080 = {"name": "0.80% Fixed", "min_quarterly_fee": 0, "tiers": [{"from": 0, "to": None, "rate": 0.0080}]}
FIXED_075 = {"name": "0.75% Fixed", "min_quarterly_fee": 0, "tiers": [{"from": 0, "to": None, "rate": 0.0075}]}
FIXED_070 = {"name": "0.70% Fixed", "min_quarterly_fee": 0, "tiers": [{"from": 0, "to": None, "rate": 0.0070}]}

ALL_SCHEDULES = {
    "gm_schedule_new": GM_SCHEDULE_NEW,
    "airline": AIRLINE_SCHEDULE,
    **{f"repricing_{k.split()[0].lower()}": v for k, v in REPRICING_SCHEDULES.items()},
    "fixed_150": FIXED_150,
    "fixed_145": FIXED_145,
    "fixed_140": FIXED_140,
    "fixed_135": FIXED_135,
    "fixed_130": FIXED_130,
    "fixed_125": FIXED_125,
    "fixed_120": FIXED_120,
    "fixed_115": FIXED_115,
    "fixed_110": FIXED_110,
    "fixed_105": FIXED_105,
    "fixed_100": FIXED_100,
    "fixed_095": FIXED_095,
    "fixed_090": FIXED_090,
    "fixed_085": FIXED_085,
    "fixed_080": FIXED_080,
    "fixed_075": FIXED_075,
    "fixed_070": FIXED_070,
}


# ─── Fee Calculation Engine ─────────────────────────────────────────────────

def calculate_tiered_fee(aum: float, schedule: dict) -> dict:
    """Calculate blended tiered fee given AUM and a schedule definition.

    Returns breakdown by tier plus summary totals.
    """
    tiers = schedule["tiers"]
    min_quarterly = schedule.get("min_quarterly_fee", 0)
    breakdown = []
    remaining = aum

    for tier in tiers:
        tier_from = tier["from"]
        tier_to = tier["to"]
        rate = tier["rate"]

        if remaining <= 0:
            breakdown.append({
                "from": tier_from,
                "to": tier_to,
                "rate": rate,
                "assets_in_tier": 0,
                "fee": 0,
            })
            continue

        if tier_to is None:
            assets_in_tier = remaining
        else:
            band_width = tier_to - tier_from + (1 if tier_from > 0 else 0)
            if tier_from == 0:
                band_width = tier_to
            assets_in_tier = min(remaining, band_width)

        fee = assets_in_tier * rate
        breakdown.append({
            "from": tier_from,
            "to": tier_to,
            "rate": rate,
            "assets_in_tier": round(assets_in_tier, 2),
            "fee": round(fee, 2),
        })
        remaining -= assets_in_tier

    total_fee = sum(t["fee"] for t in breakdown)
    annual_fee = total_fee
    quarterly_fee = annual_fee / 4

    # Apply minimum fee
    min_applied = False
    if min_quarterly > 0 and quarterly_fee < min_quarterly:
        quarterly_fee = min_quarterly
        annual_fee = quarterly_fee * 4
        min_applied = True

    effective_rate = annual_fee / aum if aum > 0 else 0

    return {
        "schedule_name": schedule["name"],
        "aum": aum,
        "breakdown": breakdown,
        "annual_fee": round(annual_fee, 2),
        "quarterly_fee": round(quarterly_fee, 2),
        "effective_rate_pct": round(effective_rate * 100, 4),
        "effective_rate_bps": round(effective_rate * 10000, 2),
        "min_fee_applied": min_applied,
        "min_quarterly_fee": min_quarterly,
    }


# ─── Database Connection (reuse app.py's pooled connection) ──────────────────

def _get_db_connection():
    """Get Synapse connection via the shared pool in app.py.

    Supports all three AUTH_METHOD modes (ActiveDirectoryInteractive,
    ServicePrincipal, SqlPassword) without any hardcoded credentials.
    """
    from app import get_database_connection
    return get_database_connection()


# ─── In-Memory Cache ─────────────────────────────────────────────────────────

_cache: dict[str, tuple[float, Any]] = {}
_CACHE_TTL_SECONDS = 1800  # 30 minutes — AUM/filters rarely change intraday


def _cache_get(key: str) -> Any | None:
    """Return cached value if still fresh, else None."""
    entry = _cache.get(key)
    if entry and (time.time() - entry[0]) < _CACHE_TTL_SECONDS:
        return entry[1]
    return None


def _cache_set(key: str, value: Any) -> None:
    _cache[key] = (time.time(), value)


def _cache_clear() -> None:
    _cache.clear()


# ─── Routes ─────────────────────────────────────────────────────────────────


@bp.route("/")
def index():
    """Serve the fee calculator SPA."""
    return send_from_directory(
        str(Path(__file__).parent / "templates"), "index.html"
    )


@bp.route("/api/cache-clear", methods=["POST"])
def clear_cache():
    """Clear all in-memory caches (forces fresh Synapse data on next request)."""
    _cache_clear()
    return jsonify({"success": True, "message": "Cache cleared"})


@bp.route("/api/schedules", methods=["GET"])
def get_schedules():
    """Return all available fee schedules and their tier definitions."""
    result = {}
    for key, sched in ALL_SCHEDULES.items():
        result[key] = {
            "name": sched["name"],
            "min_quarterly_fee": sched.get("min_quarterly_fee", 0),
            "tiers": sched["tiers"],
        }
    return jsonify({"success": True, "schedules": result})


@bp.route("/api/calculate", methods=["POST"])
def calculate():
    """Calculate fee for given AUM and schedule.

    Body: { "aum": 2000000, "schedule": "gm_schedule_new" }
    """
    data = request.get_json(force=True)
    aum = data.get("aum")
    schedule_key = data.get("schedule", "gm_schedule_new")

    if aum is None or not isinstance(aum, (int, float)) or aum < 0:
        return jsonify({"success": False, "error": "Valid 'aum' is required"}), 400

    schedule = ALL_SCHEDULES.get(schedule_key)
    if not schedule:
        return jsonify({"success": False, "error": f"Unknown schedule: {schedule_key}"}), 400

    result = calculate_tiered_fee(float(aum), schedule)
    return jsonify({"success": True, "data": result})


@bp.route("/api/calculate-all", methods=["POST"])
def calculate_all():
    """Calculate fee across ALL schedules for comparison.

    Body: { "aum": 2000000 }
    """
    data = request.get_json(force=True)
    aum = data.get("aum")

    if aum is None or not isinstance(aum, (int, float)) or aum < 0:
        return jsonify({"success": False, "error": "Valid 'aum' is required"}), 400

    results = {}
    for key, schedule in ALL_SCHEDULES.items():
        results[key] = calculate_tiered_fee(float(aum), schedule)

    return jsonify({"success": True, "data": results})


@bp.route("/api/filters", methods=["GET"])
def get_filters():
    """Return distinct values for advisor, region, and channel filters."""
    cached = _cache_get("filters")
    if cached:
        return jsonify({"success": True, "data": cached})
    try:
        conn = _get_db_connection()
        cursor = conn.cursor()

        cursor.execute(
            "SELECT DISTINCT User_ID, Name FROM [tho].[User] "
            "WHERE Department = 'Advisor' AND Name IS NOT NULL ORDER BY Name"
        )
        advisors = [{"id": str(row[0]), "name": str(row[1])} for row in cursor.fetchall()]

        cursor.execute(
            "SELECT DISTINCT Operational_Region FROM [tho].[User] "
            "WHERE Operational_Region IS NOT NULL AND Operational_Region != '' "
            "ORDER BY Operational_Region"
        )
        regions = [str(row[0]) for row in cursor.fetchall()]

        cursor.execute(
            "SELECT DISTINCT Channel_Middle FROM [tho].[Current_Household_Demographic] "
            "WHERE Channel_Middle IS NOT NULL AND Channel_Middle != '' "
            "ORDER BY Channel_Middle"
        )
        channels = [str(row[0]) for row in cursor.fetchall()]

        cursor.close()

        data = {
            "advisors": advisors,
            "regions": regions,
            "channels": channels,
        }
        _cache_set("filters", data)
        return jsonify({"success": True, "data": data})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/search", methods=["GET"])
def search_households():
    """Search households by name or advisor. Returns top 20 matches with AUM.

    Query params: q=<search term>, advisor=<user_id>, region=<name>, channel=<name>
    """
    q = request.args.get("q", "").strip()
    advisor = request.args.get("advisor", "").strip()
    region = request.args.get("region", "").strip()
    channel = request.args.get("channel", "").strip()

    if len(q) < 2 and not advisor and not region and not channel:
        return jsonify({"success": True, "results": []})

    try:
        conn = _get_db_connection()
        cursor = conn.cursor()

        # Build dynamic WHERE clause with parameterized inputs
        conditions = ["chf.Current_Client = 1"]
        params = []

        if q and len(q) >= 2:
            conditions.append("(u.Name LIKE ? OR CAST(chf.AVHHID AS VARCHAR) LIKE ?)")
            params.extend([f"%{q}%", f"%{q}%"])

        if advisor:
            conditions.append("chf.advisorid = ?")
            params.append(advisor)

        if region:
            conditions.append("u.Operational_Region = ?")
            params.append(region)

        if channel:
            conditions.append("chd.Channel_Middle = ?")
            params.append(channel)

        where_clause = " AND ".join(conditions)

        # Include demographic join only when channel filter is active
        channel_join = ""
        if channel:
            channel_join = (
                "LEFT JOIN [tho].[Current_Household_Demographic] chd "
                "ON chf.LeadId = chd.LeadId"
            )

        query = f"""
        SELECT TOP 20
            chf.AVHHID,
            chf.advisorid,
            u.Name AS advisor_name,
            u.Operational_Region AS region,
            chf.AUM,
            rf_latest.[total account value] AS current_aum
        FROM [tho].[Current_Household_Fact] chf
        LEFT JOIN [tho].[User] u ON chf.advisorid = u.User_ID
        {channel_join}
        LEFT JOIN (
            SELECT avhhid, [total account value]
            FROM [tho].[Household_Rollforward]
            WHERE reportingperiod = (
                SELECT MAX(reportingperiod) FROM [tho].[Household_Rollforward]
            )
        ) rf_latest ON chf.AVHHID = rf_latest.avhhid
        WHERE {where_clause}
        ORDER BY COALESCE(rf_latest.[total account value], chf.AUM) DESC
        """
        cursor.execute(query, params)
        columns = [desc[0] for desc in cursor.description]
        rows = cursor.fetchall()
        cursor.close()

        results = []
        for row in rows:
            row_dict = dict(zip(columns, row))
            results.append({
                "avhhid": int(row_dict["AVHHID"]),
                "advisor_name": row_dict["advisor_name"],
                "region": row_dict["region"] if row_dict.get("region") else None,
                "aum": float(row_dict["current_aum"]) if row_dict["current_aum"] else float(row_dict["AUM"] or 0),
            })

        return jsonify({"success": True, "results": results})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/household/<int:avhhid>", methods=["GET"])
def get_household(avhhid: int):
    """Get household detail with current AUM from rollforward."""
    try:
        conn = _get_db_connection()
        cursor = conn.cursor()

        query = """
        SELECT
            chf.AVHHID,
            chf.advisorid,
            u.Name AS advisor_name,
            u.Office_Location,
            chf.AUM AS fact_aum,
            rf.[total account value] AS current_aum,
            rf.reportingperiod AS aum_as_of
        FROM [tho].[Current_Household_Fact] chf
        LEFT JOIN [tho].[User] u ON chf.advisorid = u.User_ID
        LEFT JOIN (
            SELECT avhhid, [total account value], reportingperiod
            FROM [tho].[Household_Rollforward]
            WHERE reportingperiod = (
                SELECT MAX(reportingperiod) FROM [tho].[Household_Rollforward]
            )
        ) rf ON chf.AVHHID = rf.avhhid
        WHERE chf.AVHHID = ?
          AND chf.Current_Client = 1
        """
        cursor.execute(query, [avhhid])
        columns = [desc[0] for desc in cursor.description] if cursor.description else []
        row_data = cursor.fetchone()
        cursor.close()

        if not row_data:
            return jsonify({"success": False, "error": "Household not found"}), 404

        row = dict(zip(columns, row_data))
        return jsonify({
            "success": True,
            "data": {
                "avhhid": int(row["AVHHID"]),
                "advisor_name": row["advisor_name"],
                "office": row["Office_Location"],
                "current_aum": float(row["current_aum"]) if row["current_aum"] else float(row["fact_aum"] or 0),
                "aum_as_of": str(row["aum_as_of"]) if row["aum_as_of"] else None,
            },
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/upload-billing", methods=["POST"])
def upload_billing():
    """Accept a Billable_Data CSV/Excel upload.

    Aggregates by (AVHHID, Billing Definition) so each fee schedule within a
    household gets its own row.  Current AUM is sourced from
    [tav].[All_Custodian_Values] joined on Account Number, split across the
    same fee-schedule groups.
    """
    if "file" not in request.files:
        return jsonify({"success": False, "error": "No file provided"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"success": False, "error": "No file selected"}), 400

    filename = file.filename.lower()

    try:
        import pandas as pd

        if filename.endswith(".csv"):
            content = file.read().decode("utf-8-sig")
            df = pd.read_csv(io.StringIO(content))
        elif filename.endswith((".xlsx", ".xls")):
            df = pd.read_excel(file, engine="openpyxl")
        else:
            return jsonify({"success": False, "error": "Unsupported file type. Use .csv or .xlsx"}), 400

        df.columns = [c.strip() for c in df.columns]

        # Validate required columns
        required = {"AVHHID", "Billable Value", "Gross Billed Amount"}
        missing = required - set(df.columns)
        if missing:
            return jsonify({"success": False, "error": f"Missing columns: {', '.join(missing)}"}), 400

        has_account = "Account Number" in df.columns
        has_definition = "Billing Definitions" in df.columns
        has_advisor = "Advisor" in df.columns
        has_household = "Billing Group Name" in df.columns
        has_channel = "Channel" in df.columns

        # Clean currency formatting: "$1,624,879 " → 1624879.0
        def parse_money(val):
            if pd.isna(val):
                return 0.0
            s = str(val).replace("$", "").replace(",", "").strip()
            try:
                return float(s)
            except ValueError:
                return 0.0

        df["_billable"] = df["Billable Value"].apply(parse_money)
        df["_billed"] = df["Gross Billed Amount"].apply(parse_money)

        # Detect waived accounts (billing definition contains "waiv")
        if has_definition:
            df["_is_waived"] = df["Billing Definitions"].str.lower().str.contains("waiv", na=False)
            df["_group_key"] = df["Billing Definitions"].fillna("Unknown").str.strip()
        else:
            df["_is_waived"] = False
            df["_group_key"] = "Unknown"

        # --- Fetch live AUM per account from tav.All_Custodian_Values ---
        acct_aum_map: dict[str, float] = {}
        if has_account:
            cached_acct_aum = _cache_get("acct_aum_map")
            if cached_acct_aum:
                acct_aum_map = cached_acct_aum
            else:
                try:
                    conn = _get_db_connection()
                    # Use cursor directly to avoid connection sharing conflicts
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT
                            CAST([Account Number] AS VARCHAR(50)) AS acct,
                            [Total Account Value] AS tav
                        FROM [tav].[All_Custodian_Values]
                        WHERE [date] = (
                            SELECT MAX([date]) FROM [tav].[All_Custodian_Values]
                        )
                    """)
                    for row in cursor.fetchall():
                        acct = str(row[0]).strip()
                        val = row[1]
                        if val is not None and float(val) > 0:
                            acct_aum_map[acct] = float(val)
                    cursor.close()
                    _cache_set("acct_aum_map", acct_aum_map)
                except Exception as e:
                    import traceback
                    print(f"⚠️  TAV AUM lookup failed: {e}")
                    traceback.print_exc()

        # Map AUM to each row via Account Number
        if has_account and acct_aum_map:
            df["_current_aum"] = df["Account Number"].astype(str).str.strip().map(acct_aum_map).fillna(0.0)
        else:
            df["_current_aum"] = df["_billable"]  # fallback

        # --- Aggregate by (AVHHID, Billing Definition) ---
        group_cols = ["AVHHID", "_group_key"]
        agg_cols: dict = {
            "total_billable": ("_billable", "sum"),
            "quarterly_billed": ("_billed", "sum"),
            "current_aum": ("_current_aum", "sum"),
            "is_waived": ("_is_waived", "max"),  # True if any account in group is waived
            "accounts": ("AVHHID", "count"),
        }
        if has_advisor:
            agg_cols["advisor"] = ("Advisor", "first")
        if has_household:
            agg_cols["household_name"] = ("Billing Group Name", "first")
        if has_channel:
            agg_cols["channel"] = ("Channel", "first")

        agg = df.groupby(group_cols).agg(**agg_cols).reset_index()
        agg["billing_def"] = agg["_group_key"]

        # Annualise and compute current rate
        agg["annual_billed"] = agg["quarterly_billed"] * 4
        agg["current_rate_pct"] = agg.apply(
            lambda r: (r["annual_billed"] / r["total_billable"] * 100)
            if r["total_billable"] > 0 else 0.0, axis=1
        )

        # --- Build result rows ---
        households = []
        summary_stats = {
            "total_households": int(agg["AVHHID"].nunique()),
            "total_accounts": int(df["AVHHID"].count()),
            "total_billable_value": round(agg["total_billable"].sum(), 2),
            "total_non_waived_billable": round(
                agg.loc[~agg["is_waived"], "total_billable"].sum(), 2
            ),
            "total_waived_billable": round(
                agg.loc[agg["is_waived"], "total_billable"].sum(), 2
            ),
            "total_waived_accounts": int(agg.loc[agg["is_waived"], "accounts"].sum()),
            "total_annual_billed": round(agg["annual_billed"].sum(), 2),
        }

        schedule_totals_excl: dict[str, float] = {k: 0.0 for k in ALL_SCHEDULES}
        schedule_totals_incl: dict[str, float] = {k: 0.0 for k in ALL_SCHEDULES}

        for _, row in agg.iterrows():
            csv_billable = float(row["total_billable"])
            if csv_billable <= 0:
                continue

            avhhid = int(row["AVHHID"])
            current_aum = float(row["current_aum"]) if row["current_aum"] > 0 else csv_billable
            is_waived = bool(row["is_waived"])
            csv_annual = float(row["annual_billed"])
            current_rate = float(row["current_rate_pct"])

            # Recalculate current annual on current AUM so delta is apples-to-apples
            # (both current and proposed fees based on the same AUM)
            current_annual = current_aum * (current_rate / 100) if current_rate > 0 else csv_annual

            # AUM basis for proposed fee calculations — use current AUM (from TAV)
            aum_excl = current_aum
            aum_incl = current_aum

            # Auto-detect proposed schedule from AUM band
            # Waived groups or groups with no current fee → "No Change"
            no_change = False
            if is_waived or current_annual <= 0:
                no_change = True
                auto_schedule = None
            else:
                auto_schedule = _recommend_schedule_by_aum(current_aum)

            # If no recommendation ($20M+) or proposed fee < current → "No Change"
            if not no_change and auto_schedule is None:
                no_change = True
            elif not no_change:
                auto_excl = calculate_tiered_fee(aum_excl, ALL_SCHEDULES[auto_schedule])
                auto_incl = calculate_tiered_fee(aum_incl, ALL_SCHEDULES[auto_schedule])
                if auto_excl["annual_fee"] < current_annual:
                    no_change = True

            if no_change:
                hh_record_auto = {
                    "auto_schedule": "no_change",
                    "auto_schedule_name": "No Change",
                    "auto_proposed_annual": round(current_annual, 2),
                    "auto_proposed_quarterly": round(current_annual / 4, 2),
                    "auto_proposed_rate_pct": round(current_rate, 4),
                    "auto_delta": 0.0,
                    "auto_delta_pct": 0.0,
                    "auto_proposed_annual_incl": round(current_annual, 2),
                    "auto_proposed_rate_pct_incl": round(current_rate, 4),
                    "auto_delta_incl": 0.0,
                }
            else:
                ad_excl = auto_excl["annual_fee"] - current_annual
                ad_incl = auto_incl["annual_fee"] - current_annual
                hh_record_auto = {
                    "auto_schedule": auto_schedule,
                    "auto_schedule_name": ALL_SCHEDULES[auto_schedule]["name"],
                    "auto_proposed_annual": round(auto_excl["annual_fee"], 2),
                    "auto_proposed_quarterly": round(auto_excl["quarterly_fee"], 2),
                    "auto_proposed_rate_pct": round(auto_excl["effective_rate_pct"], 4),
                    "auto_delta": round(ad_excl, 2),
                    "auto_delta_pct": round(ad_excl / current_annual * 100, 2) if current_annual > 0 else 0.0,
                    "auto_proposed_annual_incl": round(auto_incl["annual_fee"], 2),
                    "auto_proposed_rate_pct_incl": round(auto_incl["effective_rate_pct"], 4),
                    "auto_delta_incl": round(ad_incl, 2),
                }

            # Calculate every schedule on both AUM bases
            proposed_excl = {}
            proposed_incl = {}
            best_key = None
            best_delta = None
            for sched_key, sched in ALL_SCHEDULES.items():
                res_excl = calculate_tiered_fee(aum_excl, sched)
                d_excl = res_excl["annual_fee"] - current_annual
                proposed_excl[sched_key] = {
                    "annual_fee": res_excl["annual_fee"],
                    "quarterly_fee": res_excl["quarterly_fee"],
                    "effective_rate_pct": res_excl["effective_rate_pct"],
                    "delta": round(d_excl, 2),
                    "delta_pct": round(d_excl / current_annual * 100, 2) if current_annual > 0 else 0.0,
                    "min_fee_applied": res_excl["min_fee_applied"],
                }
                res_incl = calculate_tiered_fee(aum_incl, sched)
                d_incl = res_incl["annual_fee"] - current_annual
                proposed_incl[sched_key] = {
                    "annual_fee": res_incl["annual_fee"],
                    "quarterly_fee": res_incl["quarterly_fee"],
                    "effective_rate_pct": res_incl["effective_rate_pct"],
                    "delta": round(d_incl, 2),
                    "delta_pct": round(d_incl / current_annual * 100, 2) if current_annual > 0 else 0.0,
                    "min_fee_applied": res_incl["min_fee_applied"],
                }
                schedule_totals_excl[sched_key] += res_excl["annual_fee"]
                schedule_totals_incl[sched_key] += res_incl["annual_fee"]
                if best_delta is None or abs(d_excl) < abs(best_delta):
                    best_delta = d_excl
                    best_key = sched_key

            hh_record: dict[str, Any] = {
                "avhhid": avhhid,
                "total_billable": round(csv_billable, 2),
                "non_waived_billable": round(0.0 if is_waived else csv_billable, 2),
                "waived_billable": round(csv_billable if is_waived else 0.0, 2),
                "waived_accounts": int(row["accounts"]) if is_waived else 0,
                "has_waived": is_waived,
                "current_aum": round(current_aum, 2),
                "accounts": int(row["accounts"]),
                "current_annual_fee": round(current_annual, 2),
                "current_quarterly_fee": round(current_annual / 4, 2),
                "current_rate_pct": round(current_rate, 4),
                "closest_schedule": best_key,
                **hh_record_auto,
                "proposed": proposed_excl,
                "proposed_incl": proposed_incl,
            }
            if has_household:
                hh_record["household_name"] = row.get("household_name", "")
            if has_advisor:
                hh_record["advisor"] = row.get("advisor", "")
            if has_definition:
                hh_record["billing_def"] = row.get("billing_def", "")
            if has_channel:
                hh_record["channel"] = row.get("channel", "")

            households.append(hh_record)

        # --- Fetch repricing campaign names by AVHHID ---
        campaign_map: dict[int, str] = {}
        avhhid_set = {h["avhhid"] for h in households}
        if avhhid_set:
            cached_campaigns = _cache_get("repricing_campaigns")
            if cached_campaigns is not None:
                campaign_map = cached_campaigns
            else:
                try:
                    conn2 = _get_db_connection()
                    cursor2 = conn2.cursor()
                    cursor2.execute("""
                        SELECT
                            chf.AVHHID,
                            STRING_AGG(cd.CampaignName, ', ') AS campaigns
                        FROM [tho].[Campaign_Demographic] cd
                        INNER JOIN [tho].[Current_Household_Fact] chf
                            ON cd.LeadId = chf.LeadId
                        WHERE cd.CampaignName LIKE '%repricing%'
                        GROUP BY chf.AVHHID
                    """)
                    for row in cursor2.fetchall():
                        campaign_map[int(row[0])] = str(row[1])
                    cursor2.close()
                    _cache_set("repricing_campaigns", campaign_map)
                except Exception as e:
                    print(f"⚠️  Campaign lookup failed: {e}")

        for hh in households:
            camp = campaign_map.get(hh["avhhid"], "")
            if camp:
                hh["campaign_name"] = camp

        # Sort by household name (groups same HH together), then AUM descending within
        households.sort(key=lambda h: (h.get("household_name", "").lower(), -h["current_aum"]))

        # Schedule-level summary (both variants)
        schedule_summary = {}
        total_current = summary_stats["total_annual_billed"]
        for sched_key, sched in ALL_SCHEDULES.items():
            tp_excl = schedule_totals_excl[sched_key]
            tp_incl = schedule_totals_incl[sched_key]
            schedule_summary[sched_key] = {
                "name": sched["name"],
                "total_annual_proposed": round(tp_excl, 2),
                "total_delta": round(tp_excl - total_current, 2),
                "delta_pct": round((tp_excl - total_current) / total_current * 100, 2) if total_current > 0 else 0.0,
                "total_annual_proposed_incl": round(tp_incl, 2),
                "total_delta_incl": round(tp_incl - total_current, 2),
                "delta_pct_incl": round((tp_incl - total_current) / total_current * 100, 2) if total_current > 0 else 0.0,
            }

        response_data = {
            "summary": summary_stats,
            "schedule_summary": schedule_summary,
            "households": households,
            "total_returned": len(households),
            "total_households": int(agg["AVHHID"].nunique()),
        }
        response_data = _json_safe(response_data)
        _cache_set("last_billing", response_data)

        return jsonify({"success": True, "data": response_data})
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to parse file: {str(e)}"}), 400


@bp.route("/api/billing-data", methods=["GET"])
def get_cached_billing():
    """Return the most recent billing upload result from cache (avoids re-upload)."""
    cached = _cache_get("last_billing")
    if cached:
        return jsonify({"success": True, "data": cached})
    return jsonify({"success": False, "error": "No billing data cached. Please upload a file."}), 404


@bp.route("/api/export-excel", methods=["POST"])
def export_excel():
    """Generate a professional Excel workbook with summary + detail on one sheet,
    advisor summary table, and interactive dropdowns.

    Body: { "households": [...], "overrides": { "rowKey": "schedKey", ... } }
    """
    import datetime
    from collections import defaultdict
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.worksheet.datavalidation import DataValidation

    data = request.get_json(force=True)
    hh_list = data.get("households", [])
    overrides = data.get("overrides", {})

    if not hh_list:
        return jsonify({"success": False, "error": "No households provided"}), 400

    wb = Workbook()

    # ─── Styles ──────────────────────────────────────────────────────────
    title_font = Font(name="Calibri", bold=True, size=16, color="1E293B")
    subtitle_font = Font(name="Calibri", size=10, color="64748B", italic=True)
    section_font = Font(name="Calibri", bold=True, size=12, color="1E293B")
    kpi_label_font = Font(name="Calibri", bold=True, size=10, color="334155")
    kpi_value_font = Font(name="Calibri", bold=True, size=11, color="1E293B")
    header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="1E293B")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    kpi_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    positive_font = Font(name="Calibri", size=10, color="16A34A")
    negative_font = Font(name="Calibri", size=10, color="DC2626")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    bottom_border = Border(bottom=Side(style="medium", color="1E293B"))

    currency_fmt = '#,##0'
    currency_dec_fmt = '$#,##0.00'
    pct_fmt = '0.00%'
    pct4_fmt = '0.0000%'

    # ─── Hidden Rate Matrix (for INDEX/MATCH formulas) ─────────────────
    ws_matrix = wb.active
    ws_matrix.title = "RateMatrix"

    sched_keys = list(ALL_SCHEDULES.keys())
    schedule_names = [ALL_SCHEDULES[k]["name"] for k in sched_keys]
    schedule_names.append("No Change")

    # Header: col A = row index, then one col per schedule
    ws_matrix.cell(row=1, column=1, value="RowIdx")
    for col_idx, name in enumerate(schedule_names, 2):
        ws_matrix.cell(row=1, column=col_idx, value=name)

    for hh_idx, hh in enumerate(hh_list):
        r = hh_idx + 2
        aum = float(hh.get("current_aum", 0))
        ws_matrix.cell(row=r, column=1, value=hh_idx + 1)
        for sched_offset, sched_key in enumerate(sched_keys):
            sched = ALL_SCHEDULES[sched_key]
            result = calculate_tiered_fee(aum, sched)
            ws_matrix.cell(row=r, column=sched_offset + 2, value=result["effective_rate_pct"] / 100)
        ws_matrix.cell(row=r, column=len(sched_keys) + 2, value=0)

    matrix_header_row = 1
    matrix_data_start = 2
    matrix_last_row = len(hh_list) + 1
    sched_col_start = 2
    matrix_last_col = len(schedule_names) + 1
    matrix_last_col_letter = get_column_letter(matrix_last_col)
    ws_matrix.sheet_state = "hidden"

    # ─── Rate Reference Sheet (visible schedule tier tables) ─────────────
    ws_ref = wb.create_sheet("Rate Reference")

    aum_fmt = '$#,##0'
    ref_row = 1
    ws_ref.merge_cells(start_row=ref_row, start_column=1, end_row=ref_row, end_column=5)
    ws_ref.cell(row=ref_row, column=1, value="Fee Schedule Reference").font = Font(name="Calibri", bold=True, size=16, color="1E293B")
    ref_row += 1
    ws_ref.cell(row=ref_row, column=1, value="Tiered and flat rate schedules used in the Fee Calculator").font = Font(name="Calibri", size=10, color="64748B", italic=True)
    ref_row += 2

    # Only show tiered schedules (multi-tier) in detail; summarize flat rates in a table
    tiered_schedules = {k: v for k, v in ALL_SCHEDULES.items() if len(v["tiers"]) > 1}
    flat_schedules = {k: v for k, v in ALL_SCHEDULES.items() if len(v["tiers"]) == 1}

    # --- Tiered schedules ---
    for sched_key, sched in tiered_schedules.items():
        # Schedule name header
        ws_ref.merge_cells(start_row=ref_row, start_column=1, end_row=ref_row, end_column=4)
        c = ws_ref.cell(row=ref_row, column=1, value=sched["name"])
        c.font = Font(name="Calibri", bold=True, size=12, color="1E293B")
        c.border = bottom_border
        ref_row += 1

        # Min fee note if applicable
        min_fee = sched.get("min_quarterly_fee", 0)
        if min_fee > 0:
            ws_ref.cell(row=ref_row, column=1, value=f"Minimum quarterly fee: ${min_fee:,.0f}").font = Font(name="Calibri", size=9, color="DC2626", italic=True)
            ref_row += 1

        # Tier table headers
        tier_headers = ["AUM From", "AUM To", "Annual Rate"]
        for col, h in enumerate(tier_headers, 1):
            c = ws_ref.cell(row=ref_row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
        ref_row += 1

        # Tier rows
        for t_idx, tier in enumerate(sched["tiers"]):
            c = ws_ref.cell(row=ref_row, column=1, value=tier["from"])
            c.number_format = aum_fmt
            c.font = data_font
            c.border = thin_border

            to_val = tier["to"] if tier["to"] is not None else "No Limit"
            if isinstance(to_val, (int, float)):
                c = ws_ref.cell(row=ref_row, column=2, value=to_val)
                c.number_format = aum_fmt
            else:
                c = ws_ref.cell(row=ref_row, column=2, value=to_val)
            c.font = data_font
            c.border = thin_border

            c = ws_ref.cell(row=ref_row, column=3, value=tier["rate"])
            c.number_format = pct_fmt
            c.font = Font(name="Calibri", size=10, bold=True, color="1D4ED8")
            c.border = thin_border
            c.alignment = Alignment(horizontal="center")

            # Alternating rows
            if t_idx % 2 == 1:
                for col in range(1, 4):
                    ws_ref.cell(row=ref_row, column=col).fill = alt_row_fill

            ref_row += 1

        ref_row += 1  # gap between schedules

    # --- Flat rate schedules table ---
    ws_ref.merge_cells(start_row=ref_row, start_column=1, end_row=ref_row, end_column=4)
    c = ws_ref.cell(row=ref_row, column=1, value="Flat Rate Schedules")
    c.font = Font(name="Calibri", bold=True, size=12, color="1E293B")
    c.border = bottom_border
    ref_row += 1

    flat_headers = ["Schedule Name", "Annual Rate"]
    for col, h in enumerate(flat_headers, 1):
        c = ws_ref.cell(row=ref_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
    ref_row += 1

    for f_idx, (sched_key, sched) in enumerate(flat_schedules.items()):
        c = ws_ref.cell(row=ref_row, column=1, value=sched["name"])
        c.font = data_font
        c.border = thin_border

        c = ws_ref.cell(row=ref_row, column=2, value=sched["tiers"][0]["rate"])
        c.number_format = pct_fmt
        c.font = Font(name="Calibri", size=10, bold=True, color="1D4ED8")
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")

        if f_idx % 2 == 1:
            for col in range(1, 3):
                ws_ref.cell(row=ref_row, column=col).fill = alt_row_fill
        ref_row += 1

    # Column widths
    ws_ref.column_dimensions["A"].width = 22
    ws_ref.column_dimensions["B"].width = 18
    ws_ref.column_dimensions["C"].width = 14
    ws_ref.column_dimensions["D"].width = 14

    # ─── Main Sheet (Summary + Detail) ───────────────────────────────────
    ws = wb.create_sheet("Fee Calculator", 0)

    # --- Title block ---
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="Allworth Fee Calculator Report").font = title_font
    row += 1
    ws.cell(row=row, column=1, value=f"Generated {datetime.datetime.now().strftime('%B %d, %Y at %I:%M %p')}").font = subtitle_font
    row += 2

    # --- KPI Summary Block ---
    ws.cell(row=row, column=1, value="PORTFOLIO SUMMARY").font = section_font
    ws.cell(row=row, column=1).border = bottom_border
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    # We'll place formulas that reference the detail table below
    # First, figure out where the detail table starts
    # Layout: summary KPIs (row 5-6), advisor table (rows 8-N), then detail table
    # We need to pre-compute advisor list for sizing
    advisors = defaultdict(lambda: {"count": 0, "aum": 0.0, "current_fees": 0.0})
    for hh in hh_list:
        adv = hh.get("advisor", "Unknown")
        advisors[adv]["count"] += 1
        advisors[adv]["aum"] += float(hh.get("current_aum", 0))
        advisors[adv]["current_fees"] += float(hh.get("current_annual_fee", 0))
    advisor_list = sorted(advisors.keys())

    # Advisor summary table starts at row+3, detail starts after that
    adv_table_start = row + 3
    adv_table_end = adv_table_start + len(advisor_list)  # header + data rows
    detail_start = adv_table_end + 3  # gap before detail
    detail_header_row = detail_start
    detail_data_start = detail_start + 1
    detail_data_end = detail_data_start + len(hh_list) - 1

    # KPI row 1: Households | Total AUM | Current Fees | Current Blended Rate
    kpi_labels_1 = ["Total Households", "Total AUM", "Current Annual Fees", "Current Blended Rate"]
    kpi_labels_2 = ["Proposed Annual Fees", "Fee Change ($)", "Fee Change (%)", "Households Changed"]

    for col, label in enumerate(kpi_labels_1, 1):
        c = ws.cell(row=row, column=col * 2 - 1, value=label)
        c.font = kpi_label_font

    row += 1
    # Values as formulas referencing detail range
    aum_range = f"E{detail_data_start}:E{detail_data_end}"
    cur_fee_range = f"F{detail_data_start}:F{detail_data_end}"
    new_fee_range = f"J{detail_data_start}:J{detail_data_end}"
    delta_range = f"K{detail_data_start}:K{detail_data_end}"

    kpi_values_1 = [
        (f"=COUNTA(A{detail_data_start}:A{detail_data_end})", currency_fmt),
        (f"=SUM({aum_range})", currency_dec_fmt),
        (f"=SUM({cur_fee_range})", currency_dec_fmt),
        (f'=IF(SUM({aum_range})=0,0,SUM({cur_fee_range})/SUM({aum_range}))', pct_fmt),
    ]
    for col, (formula, fmt) in enumerate(kpi_values_1, 1):
        c = ws.cell(row=row, column=col * 2 - 1, value=formula)
        c.font = kpi_value_font
        c.number_format = fmt
        c.fill = kpi_fill

    row += 1
    for col, label in enumerate(kpi_labels_2, 1):
        c = ws.cell(row=row, column=col * 2 - 1, value=label)
        c.font = kpi_label_font

    row += 1
    kpi_row_val2 = row
    kpi_values_2 = [
        (f"=SUM({new_fee_range})", currency_dec_fmt),
        (f"=SUM({delta_range})", currency_dec_fmt),
        (f'=IF(SUM({cur_fee_range})=0,0,SUM({delta_range})/SUM({cur_fee_range}))', pct_fmt),
        (f'=COUNTIF({delta_range},">0")+COUNTIF({delta_range},"<0")', currency_fmt),
    ]
    for col, (formula, fmt) in enumerate(kpi_values_2, 1):
        c = ws.cell(row=row, column=col * 2 - 1, value=formula)
        c.font = kpi_value_font
        c.number_format = fmt
        c.fill = kpi_fill

    # --- Advisor Summary Table ---
    row = adv_table_start - 1
    ws.cell(row=row, column=1, value="ADVISOR SUMMARY").font = section_font
    ws.cell(row=row, column=1).border = bottom_border
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

    adv_headers = ["Advisor", "Households", "Total AUM", "Current Fees",
                   "Proposed Fees", "Fee Change ($)", "Fee Change (%)", "Avg Rate"]
    row = adv_table_start
    for col, h in enumerate(adv_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    # Build advisor row references using SUMIFS against detail
    adv_col_detail = f"C{detail_data_start}:C{detail_data_end}"
    for adv_idx, adv_name in enumerate(advisor_list):
        r = row + 1 + adv_idx
        ws.cell(row=r, column=1, value=adv_name).font = data_font
        ws.cell(row=r, column=1).border = thin_border

        # Households = COUNTIF
        c = ws.cell(row=r, column=2, value=f'=COUNTIF({adv_col_detail},A{r})')
        c.number_format = '0'
        c.font = data_font
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")

        # Total AUM = SUMIF
        c = ws.cell(row=r, column=3, value=f'=SUMIF({adv_col_detail},A{r},{aum_range})')
        c.number_format = currency_dec_fmt
        c.font = data_font
        c.border = thin_border

        # Current Fees = SUMIF
        c = ws.cell(row=r, column=4, value=f'=SUMIF({adv_col_detail},A{r},{cur_fee_range})')
        c.number_format = currency_dec_fmt
        c.font = data_font
        c.border = thin_border

        # Proposed Fees = SUMIF
        c = ws.cell(row=r, column=5, value=f'=SUMIF({adv_col_detail},A{r},{new_fee_range})')
        c.number_format = currency_dec_fmt
        c.font = data_font
        c.border = thin_border

        # Fee Change ($)
        c = ws.cell(row=r, column=6, value=f'=E{r}-D{r}')
        c.number_format = currency_dec_fmt
        c.font = data_font
        c.border = thin_border

        # Fee Change (%)
        c = ws.cell(row=r, column=7, value=f'=IF(D{r}=0,0,F{r}/D{r})')
        c.number_format = pct_fmt
        c.font = data_font
        c.border = thin_border

        # Avg Rate
        c = ws.cell(row=r, column=8, value=f'=IF(C{r}=0,0,E{r}/C{r})')
        c.number_format = pct_fmt
        c.font = data_font
        c.border = thin_border

        # Alternating row color
        if adv_idx % 2 == 1:
            for col in range(1, 9):
                ws.cell(row=r, column=col).fill = alt_row_fill

    # --- Detail Table ---
    row = detail_header_row
    ws.cell(row=row - 1, column=1, value="HOUSEHOLD DETAIL").font = section_font
    ws.cell(row=row - 1, column=1).border = bottom_border
    ws.merge_cells(start_row=row - 1, start_column=1, end_row=row - 1, end_column=12)

    headers = [
        "AVHHID", "Household", "Advisor", "Billing Definition",
        "Current AUM", "Current Annual Fee", "Current Rate %",
        "AWF Schedule", "New Rate %", "New Annual Fee", "Fee Change",
        "Campaign",
    ]
    col_widths = [12, 30, 28, 20, 16, 18, 14, 24, 14, 18, 16, 20]

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border

    # Set column widths
    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Row height for header
    ws.row_dimensions[row].height = 20

    # Data validation for AWF Schedule column
    sched_start_letter = get_column_letter(sched_col_start)
    dv_list = ",".join(schedule_names)
    if len(dv_list) > 255:
        # Reference schedule names from RateMatrix header row
        dv = DataValidation(
            type="list",
            formula1=f"RateMatrix!${sched_start_letter}${matrix_header_row}:${matrix_last_col_letter}${matrix_header_row}",
            allow_blank=True,
        )
    else:
        dv = DataValidation(type="list", formula1=f'"{dv_list}"', allow_blank=True)
    dv.error = "Please select a valid fee schedule"
    dv.errorTitle = "Invalid Schedule"
    ws.add_data_validation(dv)

    # Write household rows
    for i, hh in enumerate(hh_list):
        r = detail_data_start + i
        avhhid = hh.get("avhhid", "")
        row_key = f"{avhhid}-{hh.get('billing_def', '')}"
        current_aum = float(hh.get("current_aum", 0))
        current_annual = float(hh.get("current_annual_fee", 0))
        current_rate = float(hh.get("current_rate_pct", 0)) / 100

        # AWF override
        override_key = overrides.get(row_key, "")
        if override_key == "no_change":
            awf_name = "No Change"
        elif override_key:
            awf_name = ALL_SCHEDULES.get(override_key, {}).get("name", "")
        else:
            awf_name = ""

        ws.cell(row=r, column=1, value=avhhid).font = data_font
        ws.cell(row=r, column=2, value=hh.get("household_name", "")).font = data_font
        ws.cell(row=r, column=3, value=hh.get("advisor", "")).font = data_font
        ws.cell(row=r, column=4, value=hh.get("billing_def", "")).font = data_font

        c = ws.cell(row=r, column=5, value=current_aum)
        c.number_format = currency_dec_fmt
        c.font = data_font

        c = ws.cell(row=r, column=6, value=current_annual)
        c.number_format = currency_dec_fmt
        c.font = data_font

        c = ws.cell(row=r, column=7, value=current_rate)
        c.number_format = pct4_fmt
        c.font = data_font

        # Dropdown
        c_sched = ws.cell(row=r, column=8, value=awf_name)
        c_sched.font = Font(name="Calibri", size=10, bold=True, color="1D4ED8")
        dv.add(c_sched)

        # New Rate = INDEX/MATCH from RateMatrix (AUM-specific tiered rates)
        # Row in matrix = this household's index (i+1), matched via col A
        # Col in matrix = schedule name from dropdown, matched via header row
        hh_matrix_idx = i + 1  # 1-based index stored in RateMatrix col A
        rate_col = "H"
        sched_start_letter = get_column_letter(sched_col_start)
        index_match = (
            f'=IF({rate_col}{r}="","",'
            f'IF({rate_col}{r}="No Change",G{r},'
            f"INDEX(RateMatrix!${sched_start_letter}${matrix_data_start}:${matrix_last_col_letter}${matrix_last_row},"
            f"MATCH({hh_matrix_idx},RateMatrix!$A${matrix_data_start}:$A${matrix_last_row},0),"
            f"MATCH({rate_col}{r},RateMatrix!${sched_start_letter}${matrix_header_row}:${matrix_last_col_letter}${matrix_header_row},0))))"
        )
        c = ws.cell(row=r, column=9, value=index_match)
        c.number_format = pct4_fmt
        c.font = data_font

        # New Annual Fee (blank when no schedule selected)
        c = ws.cell(row=r, column=10, value=f'=IF(H{r}="","",E{r}*I{r})')
        c.number_format = currency_dec_fmt
        c.font = data_font

        # Fee Change (blank when no schedule selected)
        c = ws.cell(row=r, column=11, value=f'=IF(H{r}="","",J{r}-F{r})')
        c.number_format = currency_dec_fmt
        c.font = data_font

        ws.cell(row=r, column=12, value=hh.get("campaign_name", "")).font = data_font

        # Apply borders and alternating rows
        for col in range(1, 13):
            ws.cell(row=r, column=col).border = thin_border
            ws.cell(row=r, column=col).alignment = Alignment(vertical="center")
        if i % 2 == 1:
            for col in range(1, 13):
                ws.cell(row=r, column=col).fill = alt_row_fill

    # Freeze panes at detail header
    ws.freeze_panes = f"A{detail_data_start}"

    # ─── Write to bytes and return ───────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Build filename: AdvisorName_Fee_Repricing-Date (or "Multiple_Advisors" if >1)
    unique_advisors = list({hh.get("advisor", "") for hh in hh_list if hh.get("advisor")})
    if len(unique_advisors) == 1:
        adv_name = unique_advisors[0].replace(" ", "_").replace(",", "")
    else:
        adv_name = "Multiple_Advisors"
    file_date = datetime.date.today().strftime("%m-%d-%Y")
    filename = f"{adv_name}_Fee_Repricing-{file_date}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def _recommend_schedule_by_aum(aum: float) -> str | None:
    """Recommend a fee schedule based on the client's AUM band.

    Returns the schedule key, or None if AUM > $20M (no recommendation).
    """
    if aum <= 300_000:
        return "repricing_silver"
    elif aum <= 1_000_000:
        return "repricing_gold"
    elif aum <= 2_000_000:
        return "repricing_platinum"
    elif aum <= 5_000_000:
        return "repricing_elite"
    elif aum <= 7_500_000:
        return "fixed_080"
    elif aum <= 10_000_000:
        return "fixed_075"
    elif aum <= 20_000_000:
        return "fixed_070"
    else:
        return None  # $20M+ — leave blank (no recommendation)

